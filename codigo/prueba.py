import tkinter as tk
from tkinter import filedialog, messagebox
import os, shutil, re
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

class PDFPage:
    def __init__(self, text=""):
        self.text = text
        self.tables = []

class PDFDocument:
    def __init__(self, pages=None):
        self.pages = pages if pages else []

def ajustar_columnas_filas(ws):
    for col in range(1, 4):
        max_length = max((len(str(cell.value)) for cell in ws[get_column_letter(col)]), default=0)
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2
    for row in ws.iter_rows():
        max_height = max((len(str(cell.value).split('\n')) for cell in row), default=1)
        ws.row_dimensions[row[0].row].height = max_height * 14

def agregar_pdf_a_excel(pdf_documento, ruta_excel, nombre_hoja):
    wb = load_workbook(ruta_excel)
    nombre_hoja = nombre_hoja[:31]  # Truncar nombre si excede 31 caracteres
    ws = wb.create_sheet(title=nombre_hoja)
    fuente_titulo = Font(name='Arial', size=12, bold=True)
    fuente_contenido = Font(name='Arial', size=10)
    alineacion_izquierda = Alignment(horizontal='left', vertical='top', wrap_text=True)

    fila_idx = 1
    for pagina in pdf_documento.pages:
        for linea in pagina.text.split("\n"):
            if linea.strip():
                celda = ws.cell(row=fila_idx, column=1, value=linea.strip())
                ws.merge_cells(start_row=fila_idx, start_column=1, end_row=fila_idx, end_column=3)
                celda.font = fuente_titulo if fila_idx == 1 else fuente_contenido
                celda.alignment = alineacion_izquierda
                if linea.startswith(("●", "•", "○")) or re.match(r"^\d+\.", linea.strip()):
                    celda.font = Font(name='Arial', size=10, bold=True)
                fila_idx += 1
    ajustar_columnas_filas(ws)
    wb.save(ruta_excel)

def agregar_excel_a_excel(origen, destino, nombre_hoja):
    wb_origen, wb_destino = load_workbook(origen), load_workbook(destino)
    ws_origen, ws_destino = wb_origen.active, wb_destino.create_sheet(title=nombre_hoja[:31])  # Limitar nombre a 31 caracteres
    for fila in ws_origen.iter_rows():
        for celda in fila:
            nueva_celda = ws_destino.cell(row=celda.row, column=celda.column, value=celda.value)
            if celda.font:  # Evitar error con StyleProxy para las fuentes
                nueva_celda.font = Font(
                    name=celda.font.name, size=celda.font.size, bold=celda.font.bold,
                    italic=celda.font.italic, vertAlign=celda.font.vertAlign,
                    underline=celda.font.underline, strike=celda.font.strike, color=celda.font.color
                )
            if celda.alignment:  # Evitar error con StyleProxy para la alineación
                nueva_celda.alignment = Alignment(
                    horizontal=celda.alignment.horizontal,
                    vertical=celda.alignment.vertical,
                    text_rotation=celda.alignment.textRotation,
                    wrap_text=celda.alignment.wrapText,
                    shrink_to_fit=celda.alignment.shrinkToFit,
                    indent=celda.alignment.indent
                )
            else:
                nueva_celda.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ajustar_columnas_filas(ws_destino)
    wb_destino.save(destino)


def eliminar_pie_de_pagina(ruta_excel):
    wb = load_workbook(ruta_excel)
    for hoja in wb:
        for fila in hoja.iter_rows():
            for celda in fila:
                if isinstance(celda.value, str) and re.sub(r'[^\w\s]', '', celda.value) == "GFPI135V01":
                    celda.value = None
    wb.save(ruta_excel)

def leer_pdf(ruta_pdf):
    doc = fitz.open(ruta_pdf)
    return [PDFPage(text=doc.load_page(i).get_text()) for i in range(len(doc))]

# Clases y funciones como están...

def procesar_archivos():
    global ruta_plantilla, ruta_guardado
    if not ruta_plantilla:
        return messagebox.showerror("Error", "Seleccione una plantilla")
    if not ruta_guardado:
        return messagebox.showerror("Error", "Seleccione una ubicación para guardar el archivo")

    pdfs, excels = [f for f in archivos if f.endswith(".pdf")], [f for f in archivos if f.endswith(".xlsx")]
    if not pdfs and not excels:
        return messagebox.showerror("Error", "Seleccione archivos PDF o Excel")

    copia_excel = f"{ruta_guardado}.xlsx"  # Usar la ruta seleccionada por el usuario para guardar
    shutil.copyfile(ruta_plantilla, copia_excel)

    for pdf in pdfs:
        agregar_pdf_a_excel(PDFDocument(pages=leer_pdf(pdf)), copia_excel, os.path.basename(pdf))
    for excel in excels:
        agregar_excel_a_excel(excel, copia_excel, os.path.basename(excel))

    eliminar_pie_de_pagina(copia_excel)
    messagebox.showinfo("Éxito", f"Procesamiento completado en '{copia_excel}'")

def actualizar_listbox():
    """Actualiza el Listbox con los archivos seleccionados."""
    listbox_archivos.delete(0, tk.END)  # Limpiar el Listbox antes de actualizar
    for archivo in archivos:
        listbox_archivos.insert(tk.END, archivo)

def seleccionar_archivos():
    """Función para seleccionar archivos y actualizar el Listbox."""
    global archivos
    archivos_seleccionados = filedialog.askopenfilenames(filetypes=[("PDF y Excel", "*.pdf *.xlsx")])
    if archivos_seleccionados:
        archivos.extend(archivos_seleccionados)
        actualizar_listbox()

def crear_gui():
    global listbox_archivos, ruta_plantilla, archivos, ruta_guardado
    ruta_plantilla, archivos, ruta_guardado = None, [], None
    root = tk.Tk()
    root.title("Procesador de PDF y Excel")
    
    tk.Label(root, text="Seleccione archivos PDF y Excel").pack(pady=10)
    tk.Button(root, text="Seleccionar archivos", command=seleccionar_archivos).pack(pady=5)
    listbox_archivos = tk.Listbox(root, width=50, height=10)
    listbox_archivos.pack(pady=10)

    etiqueta_plantilla = tk.Label(root, text="Plantilla no seleccionada")
    etiqueta_plantilla.pack(pady=10)
    def seleccionar_plantilla():
        global ruta_plantilla
        ruta_plantilla = filedialog.askopenfilename()
        if ruta_plantilla:
            etiqueta_plantilla.config(text=f"Plantilla: {os.path.basename(ruta_plantilla)}")
    
    tk.Button(root, text="Seleccionar plantilla", command=seleccionar_plantilla).pack(pady=5)

    etiqueta_guardado = tk.Label(root, text="Ubicación de guardado no seleccionada")
    etiqueta_guardado.pack(pady=10)
    def seleccionar_guardado():
        global ruta_guardado
        ruta_guardado = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if ruta_guardado:
            etiqueta_guardado.config(text=f"Guardado en: {os.path.basename(ruta_guardado)}")
    
    tk.Button(root, text="Seleccionar ubicación para guardar", command=seleccionar_guardado).pack(pady=5)

    tk.Button(root, text="Procesar archivos", command=procesar_archivos).pack(pady=5)
    root.mainloop()

crear_gui()
